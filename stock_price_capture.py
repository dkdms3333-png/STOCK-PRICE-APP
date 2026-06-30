import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date, timedelta
import requests
from bs4 import BeautifulSoup
import re
import time
import io
import os
import subprocess
import zipfile
import FinanceDataReader as fdr
from PIL import Image


# ── Playwright Chromium 자동 설치 (Streamlit Cloud) ──────────
@st.cache_resource
def ensure_playwright_browser():
    """Streamlit Cloud에서 Chromium 자동 설치. 결과 메시지 반환."""
    msgs = []
    try:
        result = subprocess.run(
            ["playwright", "install", "chromium"],
            capture_output=True, timeout=300, text=True,
        )
        msgs.append(f"install rc={result.returncode}")
        if result.stderr:
            msgs.append(f"stderr: {result.stderr[-300:]}")
    except Exception as e:
        msgs.append(f"install exception: {e}")

    # 실제로 동작하는지 테스트
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            b = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            b.close()
        msgs.append("launch OK")
        return True, msgs
    except Exception as e:
        msgs.append(f"launch failed: {e}")
        return False, msgs

st.set_page_config(page_title="종가 캡처", layout="centered")
st.title("코스닥/코스피 종가 조회")

# ── 종목코드 조회 ─────────────────────────────────────────────
def clean_stock_name(name: str) -> str:
    return re.sub(r'[㈜㈔\s]', '', name).strip()


@st.cache_data(ttl=3600 * 12, show_spinner="전종목 코드 로딩 중 (최초 1회)...")
def get_code_map() -> dict:
    """FinanceDataReader로 코스피+코스닥 전종목 코드맵 반환."""
    code_map = {}
    for market in ["KOSPI", "KOSDAQ"]:
        try:
            df = fdr.StockListing(market)
            for _, row in df.iterrows():
                name = str(row.get("Name", "")).strip()
                code = str(row.get("Code", "")).strip()
                if name and code:
                    code_map[name] = code
                    code_map[clean_stock_name(name)] = code
        except Exception:
            pass
    return code_map


def get_stock_code(name: str, code_map: dict) -> str | None:
    name = name.strip()
    return code_map.get(name) or code_map.get(clean_stock_name(name))


# ── 종가 조회 ────────────────────────────────────────────────
def fetch_closing_price(code: str, target_date: date) -> int | None:
    headers = {"User-Agent": "Mozilla/5.0"}
    date_str = target_date.strftime("%Y.%m.%d")
    today = date.today()
    trading_days = (today - target_date).days * 5 // 7
    start_page = max(1, trading_days // 10 - 2)

    for page in range(start_page, start_page + 15):
        url = f"https://finance.naver.com/item/sise_day.nhn?code={code}&page={page}"
        res = requests.get(url, headers=headers, timeout=10)
        res.encoding = "euc-kr"
        soup = BeautifulSoup(res.text, "html.parser")
        rows = soup.select("table.type2 tr")
        for row in rows:
            tds = row.select("td")
            if len(tds) < 2:
                continue
            row_date = tds[0].text.strip()
            if row_date == date_str:
                price_text = tds[1].text.strip().replace(",", "")
                try:
                    return int(price_text)
                except ValueError:
                    return None
            if row_date and row_date < date_str:
                break
    return None


def find_prev_trading_day(code: str, target_date: date) -> tuple[date | None, int | None]:
    for delta in range(0, 6):
        check_date = target_date - timedelta(days=delta)
        price = fetch_closing_price(code, check_date)
        if price:
            return check_date, price
    return None, None


# ── 스크린샷 (네이버 시세 페이지) ────────────────────────────
def find_page_for_date(code: str, target_date: date) -> int:
    headers = {"User-Agent": "Mozilla/5.0"}
    date_str = target_date.strftime("%Y.%m.%d")
    today = date.today()
    trading_days = (today - target_date).days * 5 // 7
    start_page = max(1, trading_days // 10 - 2)
    for page in range(start_page, start_page + 10):
        url = f"https://finance.naver.com/item/sise_day.nhn?code={code}&page={page}"
        res = requests.get(url, headers=headers, timeout=10)
        res.encoding = "euc-kr"
        soup = BeautifulSoup(res.text, "html.parser")
        for row in soup.select("table.type2 tr"):
            tds = row.select("td")
            if tds and tds[0].text.strip() == date_str:
                return page
            if tds and tds[0].text.strip() and tds[0].text.strip() < date_str:
                break
    return start_page


def capture_naver_chart(code: str, actual_date: date) -> tuple[dict, str]:
    """sise.naver 페이지를 그대로 열고, 일별시세 iframe만 기준일 페이지로 넘김.
    페이지 구조/내용 수정 없음 — iframe의 페이지 번호만 네비게이션.
    """
    from playwright.sync_api import sync_playwright
    target_page = find_page_for_date(code, actual_date)
    sise_url = f"https://finance.naver.com/item/sise.naver?code={code}"
    day_url  = f"https://finance.naver.com/item/sise_day.naver?code={code}&page={target_page}"
    date_str = actual_date.strftime("%Y.%m.%d")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            page = browser.new_page(viewport={"width": 1280, "height": 1600})
            page.goto(sise_url, wait_until="domcontentloaded", timeout=20000)
            time.sleep(2)

            # day iframe의 src를 기준일 페이지로 변경 (사용자가 페이지 번호 클릭한 것과 동일)
            page.evaluate(f"""
                var f = document.querySelector('iframe[name=day]');
                if (f) f.src = '{day_url}';
            """)

            # iframe이 기준일 텍스트를 포함할 때까지 대기 (최대 8초)
            deadline = time.time() + 8
            ok = False
            while time.time() < deadline:
                for f in page.frames:
                    if f.name == "day":
                        try:
                            txt = f.evaluate("document.body.innerText")
                            if date_str in txt:
                                ok = True
                                break
                        except Exception:
                            pass
                if ok:
                    break
                time.sleep(0.5)
            time.sleep(0.5)

            # 일별시세 iframe의 bottom 좌표를 측정해서 그 지점까지만 캡처
            bbox = page.evaluate("""
                () => {
                    const f = document.querySelector('iframe[name=day]');
                    if (!f) return null;
                    const r = f.getBoundingClientRect();
                    return {
                        bottom: Math.ceil(r.bottom + window.scrollY),
                        right: Math.ceil(r.right + window.scrollX)
                    };
                }
            """)
            if bbox:
                clip_h = bbox["bottom"] + 20
                clip_w = min(1280, max(bbox["right"] + 20, 900))
                img = page.screenshot(clip={"x": 0, "y": 0, "width": clip_w, "height": clip_h})
            else:
                img = page.screenshot(full_page=True)

            browser.close()
            err = "" if ok else "기준일 로드 확인 실패 (캡처는 진행됨)"
            return {"전체": img}, err
    except Exception as e:
        return {}, str(e)[:200]


# ── 엑셀 생성 (다운로드용 메모리 버퍼) ──────────────────────
def save_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "종가현황"

    headers = ["펀드명", "종목명", "기준일", "실제 조회일", "종가(원)"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["fund"]).border = border
        ws.cell(r, 2, row["name"]).border = border
        ws.cell(r, 3, row["ref_date"]).border = border
        actual = ws.cell(r, 4, row["actual_date"])
        actual.border = border
        if row["ref_date"] != row["actual_date"]:
            actual.font = Font(color="FF0000")
        price_cell = ws.cell(r, 5, row["price"])
        price_cell.number_format = "#,##0"
        price_cell.alignment = Alignment(horizontal="right")
        price_cell.border = border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ───────────────────────────────────────────────────────
st.markdown("### 1. 엑셀 업로드")
st.caption("**펀드명**, **종목명** 두 열 포함. 헤더 행 필수.")

uploaded = st.file_uploader("종목 목록 엑셀 (.xlsx)", type=["xlsx"])
target_date = st.date_input("기준일", value=date.today())
capture_enabled = st.checkbox("📸 네이버 시세 화면 캡처 (분기결산 증빙용)", value=False,
                              help="체크 시 각 종목별 네이버 화면을 PNG로 저장. 처리 시간이 길어집니다.")
st.caption("⚠️ 기준일이 휴장일이면 직전 영업일 종가로 자동 대체됩니다. (엑셀에 빨간색 표시)")

if uploaded:
    try:
        df = pd.read_excel(uploaded)
        df.columns = df.columns.str.strip()

        col_fund = next((c for c in df.columns if "펀드" in str(c)), None)
        col_name = next((c for c in df.columns if "종목" in str(c)), None)

        if not col_fund or not col_name:
            st.error("엑셀에 '펀드명'과 '종목명' 열이 필요합니다.")
        else:
            df = df[[col_fund, col_name]].dropna()
            st.dataframe(df, use_container_width=True)
            st.info(f"총 {len(df)}개 종목")

            if st.button("종가 조회 시작", type="primary"):
                code_map = get_code_map()
                if not code_map:
                    st.error("종목코드 로딩 실패. 잠시 후 다시 시도해주세요.")
                    st.stop()
                init_msgs = []
                if capture_enabled:
                    with st.spinner("브라우저 초기화 중 (최초 1회 1~2분)..."):
                        ok, init_msgs = ensure_playwright_browser()
                    if not ok:
                        st.error("브라우저 초기화 실패. 캡처 없이 종가 조회만 진행합니다.")
                        capture_enabled = False
                results = []
                errors = []
                captures: dict[str, bytes] = {}
                progress = st.progress(0)
                status = st.empty()

                for i, (_, row) in enumerate(df.iterrows()):
                    fund_name = str(row[col_fund]).strip()
                    stock_name = str(row[col_name]).strip()
                    status.text(f"처리 중: {stock_name} ({i+1}/{len(df)})")

                    code = get_stock_code(stock_name, code_map)
                    if not code:
                        errors.append(f"{stock_name}: 종목코드 없음")
                        results.append({
                            "fund": fund_name, "name": stock_name,
                            "ref_date": target_date.strftime("%Y-%m-%d"),
                            "actual_date": "-", "price": "코드 없음",
                        })
                        progress.progress((i + 1) / len(df))
                        continue

                    actual_date, price = find_prev_trading_day(code, target_date)

                    if not price:
                        errors.append(f"{stock_name}: 종가 조회 실패")
                        results.append({
                            "fund": fund_name, "name": stock_name,
                            "ref_date": target_date.strftime("%Y-%m-%d"),
                            "actual_date": "-", "price": "조회 실패",
                        })
                        progress.progress((i + 1) / len(df))
                        continue

                    results.append({
                        "fund": fund_name, "name": stock_name,
                        "ref_date": target_date.strftime("%Y-%m-%d"),
                        "actual_date": actual_date.strftime("%Y-%m-%d"),
                        "price": price,
                    })

                    if capture_enabled:
                        status.text(f"캡처 중: {stock_name} ({i+1}/{len(df)})")
                        imgs, err = capture_naver_chart(code, actual_date)
                        date_label = target_date.strftime("%Y%m%d")
                        for suffix, data in imgs.items():
                            captures[f"{stock_name}_{date_label}_{suffix}.png"] = data
                        if not imgs:
                            errors.append(f"{stock_name}: 캡처 실패 — {err}")
                        elif err:
                            errors.append(f"{stock_name}: 일부 캡처 실패 — {err}")

                    progress.progress((i + 1) / len(df))

                status.text("완료!")

                excel_rows = [r for r in results if isinstance(r["price"], int)]
                date_label = target_date.strftime("%Y%m%d")
                excel_bytes = save_excel(excel_rows) if excel_rows else None

                zip_bytes = None
                if captures:
                    zip_buf = io.BytesIO()
                    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                        for fname, data in captures.items():
                            zf.writestr(fname, data)
                    zip_bytes = zip_buf.getvalue()

                # session_state에 저장 → 다운로드 클릭 후에도 결과 유지
                st.session_state["last_results"] = results
                st.session_state["last_errors"] = errors
                st.session_state["last_excel"] = excel_bytes
                st.session_state["last_zip"] = zip_bytes
                st.session_state["last_zip_count"] = len(captures)
                st.session_state["last_date_label"] = date_label
                st.session_state["last_init_msgs"] = init_msgs

    except Exception as e:
        st.error(f"오류: {e}")
        import traceback
        st.code(traceback.format_exc())

# ── 결과 표시 (session_state 기반, 다운로드해도 유지됨) ──────
if "last_results" in st.session_state:
    st.markdown("---")
    st.markdown("### 결과")
    result_df = pd.DataFrame(st.session_state["last_results"])
    result_df.columns = ["펀드명", "종목명", "기준일", "실제조회일", "종가(원)"]
    st.dataframe(result_df, use_container_width=True)

    col_a, col_b = st.columns(2)
    date_label = st.session_state["last_date_label"]
    if st.session_state.get("last_excel"):
        with col_a:
            st.download_button(
                label="📥 엑셀 다운로드",
                data=st.session_state["last_excel"],
                file_name=f"종가현황_{date_label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel",
            )
    if st.session_state.get("last_zip"):
        with col_b:
            st.download_button(
                label=f"📸 캡처 ZIP 다운로드 ({st.session_state['last_zip_count']}장)",
                data=st.session_state["last_zip"],
                file_name=f"종가캡처_{date_label}.zip",
                mime="application/zip",
                key="dl_zip",
            )

    if st.session_state.get("last_init_msgs"):
        with st.expander("브라우저 초기화 로그"):
            for m in st.session_state["last_init_msgs"]:
                st.text(m)

    if st.session_state.get("last_errors"):
        with st.expander("오류 목록"):
            for e in st.session_state["last_errors"]:
                st.warning(e)
